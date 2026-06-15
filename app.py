import os
import io
import datetime
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from werkzeug.security import check_password_hash, generate_password_hash
from database import get_db_connection, init_db

# Initialize database
init_db()

app = Flask(__name__)
app.secret_key = 'stockflow_secret_key_for_production_and_testing'

# Context processor to make active route check easier in templates
@app.context_processor
def inject_now():
    return {'now': datetime.datetime.utcnow()}

# Authentication decorator
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('logged_in'):
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE username = ?", (username,))
        user = cursor.fetchone()
        conn.close()
        
        if user and check_password_hash(user['password'], password):
            session['logged_in'] = True
            session['username'] = username
            flash('Login realizado com sucesso!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Usuário ou senha incorretos.', 'danger')
            
    return render_template('login.html')

@app.route('/user/add', methods=['POST'])
@login_required
def add_user():
    username = request.form.get('username', '').strip()
    password = request.form.get('password')
    confirm_password = request.form.get('confirm_password')
    
    if not username or not password or not confirm_password:
        flash('Por favor, preencha todos os campos para o novo usuário.', 'danger')
        return redirect(url_for('usuarios'))
        
    if password != confirm_password:
        flash('As senhas não coincidem para o novo usuário.', 'danger')
        return redirect(url_for('usuarios'))
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE username = ?", (username,))
    existing_user = cursor.fetchone()
    
    if existing_user:
        flash('Este usuário já existe.', 'danger')
        conn.close()
        return redirect(url_for('usuarios'))
        
    hashed_password = generate_password_hash(password)
    cursor.execute("INSERT INTO users (username, password) VALUES (?, ?)", (username, hashed_password))
    conn.commit()
    conn.close()
    
    flash(f'Usuário {username} cadastrado com sucesso!', 'success')
    return redirect(url_for('usuarios'))

@app.route('/user/edit/<int:user_id>', methods=['POST'])
@login_required
def edit_user(user_id):
    username = request.form.get('username', '').strip()
    password = request.form.get('password')
    confirm_password = request.form.get('confirm_password')
    
    if not username:
        flash('O nome de usuário não pode ser vazio.', 'danger')
        return redirect(url_for('usuarios'))
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Check if new username conflicts with another user
    cursor.execute("SELECT * FROM users WHERE username = ? AND id != ?", (username, user_id))
    existing_user = cursor.fetchone()
    if existing_user:
        flash('Já existe outro usuário com este nome.', 'danger')
        conn.close()
        return redirect(url_for('usuarios'))
        
    # Get current username of the edited user to see if it's the current user
    cursor.execute("SELECT username FROM users WHERE id = ?", (user_id,))
    current_user_row = cursor.fetchone()
    is_self = (current_user_row and current_user_row['username'] == session.get('username'))
    
    if password:
        if password != confirm_password:
            flash('As senhas não coincidem.', 'danger')
            conn.close()
            return redirect(url_for('usuarios'))
        hashed_password = generate_password_hash(password)
        cursor.execute("UPDATE users SET username = ?, password = ? WHERE id = ?", (username, hashed_password, user_id))
    else:
        cursor.execute("UPDATE users SET username = ? WHERE id = ?", (username, user_id))
        
    conn.commit()
    conn.close()
    
    # If the user edited themselves, update the session
    if is_self:
        session['username'] = username
        
    flash('Usuário atualizado com sucesso!', 'success')
    return redirect(url_for('usuarios'))

@app.route('/user/delete/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Check count of users
    cursor.execute("SELECT COUNT(*) as qty FROM users")
    count_row = cursor.fetchone()
    if count_row['qty'] <= 1:
        flash('Não é possível excluir o único usuário do sistema.', 'danger')
        conn.close()
        return redirect(url_for('usuarios'))
        
    # Check if deleting self
    cursor.execute("SELECT username FROM users WHERE id = ?", (user_id,))
    user_row = cursor.fetchone()
    if user_row and user_row['username'] == session.get('username'):
        flash('Você não pode excluir o seu próprio usuário enquanto estiver logado.', 'danger')
        conn.close()
        return redirect(url_for('usuarios'))
        
    cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    
    flash('Usuário excluído com sucesso.', 'success')
    return redirect(url_for('usuarios'))

@app.route('/logout')
def logout():
    session.clear()
    flash('Você saiu do sistema.', 'success')
    return redirect(url_for('login'))

@app.route('/usuarios')
@login_required
def usuarios():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, username FROM users ORDER BY username")
    users = cursor.fetchall()
    conn.close()
    return render_template('usuarios.html', users=users)

@app.route('/')
@login_required
def dashboard():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get all products and their status
    cursor.execute("""
        SELECT ssd_type, brand, size, quantity, total_exited 
        FROM products 
        ORDER BY ssd_type, brand, size
    """)
    products = cursor.fetchall()
    
    # Calculate stats
    cursor.execute("SELECT SUM(quantity) as total_stock, SUM(total_exited) as total_exits FROM products")
    stats = cursor.fetchone()
    
    # Detailed count for chart/breakdown
    cursor.execute("SELECT ssd_type, SUM(quantity) as qty FROM products GROUP BY ssd_type")
    type_breakdown = {row['ssd_type']: row['qty'] or 0 for row in cursor.fetchall()}
    
    conn.close()
    
    total_stock = stats['total_stock'] if stats and stats['total_stock'] is not None else 0
    total_exits = stats['total_exits'] if stats and stats['total_exits'] is not None else 0
    
    sata_qty = type_breakdown.get('SATA', 0)
    nvme_qty = type_breakdown.get('NVMe', 0)
    
    # Detect low stock alert (e.g. quantity <= 3)
    low_stock_items = [p for p in products if p['quantity'] <= 3]
    
    return render_template('dashboard.html', 
                           products=products, 
                           total_stock=total_stock, 
                           total_exits=total_exits,
                           sata_qty=sata_qty,
                           nvme_qty=nvme_qty,
                           low_stock_count=len(low_stock_items))

@app.route('/entradas', methods=['GET', 'POST'])
@login_required
def entradas():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        entry_date = request.form['entry_date']
        ssd_type = request.form['ssd_type']
        brand = request.form['brand'].strip()
        size = request.form['size'].strip() or None
        quantity = int(request.form['quantity'])
        supplier = request.form['supplier'].strip()
        price = float(request.form['price'])
        
        # Save entry record
        cursor.execute("""
            INSERT INTO entries (entry_date, ssd_type, brand, size, quantity, supplier, price)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (entry_date, ssd_type, brand, size, quantity, supplier, price))
        
        # Update or create product aggregate
        cursor.execute("""
            SELECT * FROM products WHERE ssd_type = ? AND brand = ? AND (size = ? OR (size IS NULL AND ? IS NULL))
        """, (ssd_type, brand, size, size))
        product = cursor.fetchone()
        
        if product:
            cursor.execute("""
                UPDATE products 
                SET quantity = quantity + ? 
                WHERE id = ?
            """, (quantity, product['id']))
        else:
            cursor.execute("""
                INSERT INTO products (ssd_type, brand, size, quantity, total_exited)
                VALUES (?, ?, ?, ?, 0)
            """, (ssd_type, brand, size, quantity))
            
        conn.commit()
        flash('Entrada de produto registrada com sucesso!', 'success')
        conn.close()
        return redirect(url_for('entradas'))
        
    # Get all entries list
    cursor.execute("SELECT * FROM entries ORDER BY entry_date DESC, id DESC")
    entries_list = cursor.fetchall()
    conn.close()
    
    return render_template('entradas.html', entries=entries_list)

@app.route('/saidas', methods=['GET', 'POST'])
@login_required
def saidas():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        exit_date = request.form['exit_date']
        ssd_type = request.form['ssd_type']
        brand = request.form['brand'].strip()
        size = request.form['size'].strip() or None
        quantity = int(request.form['quantity'])
        supplier = request.form['supplier'].strip() # user requested supplier on exit as well
        client = request.form['client'].strip()
        
        # Check stock availability
        cursor.execute("""
            SELECT * FROM products 
            WHERE ssd_type = ? AND brand = ? AND (size = ? OR (size IS NULL AND ? IS NULL))
        """, (ssd_type, brand, size, size))
        product = cursor.fetchone()
        
        if not product or product['quantity'] < quantity:
            avail = product['quantity'] if product else 0
            flash(f'Quantidade insuficiente no estoque. Disponível: {avail} unidades.', 'danger')
            conn.close()
            return redirect(url_for('saidas'))
            
        # Register exit record
        cursor.execute("""
            INSERT INTO exits (exit_date, ssd_type, brand, size, quantity, supplier, client)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (exit_date, ssd_type, brand, size, quantity, supplier, client))
        
        # Update product aggregate
        cursor.execute("""
            UPDATE products 
            SET quantity = quantity - ?, total_exited = total_exited + ? 
            WHERE id = ?
        """, (quantity, quantity, product['id']))
        
        conn.commit()
        flash('Saída de produto registrada com sucesso!', 'success')
        conn.close()
        return redirect(url_for('saidas'))
        
    # Get all exits list
    cursor.execute("SELECT * FROM exits ORDER BY exit_date DESC, id DESC")
    exits_list = cursor.fetchall()
    
    # Get active products for select autocomplete helpers
    cursor.execute("SELECT ssd_type, brand, size, quantity FROM products WHERE quantity > 0 ORDER BY brand, size")
    available_products = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    
    return render_template('saidas.html', exits=exits_list, available_products=available_products)

# Edit Entry
@app.route('/entradas/edit/<int:entry_id>', endpoint='edit_entry', methods=['GET', 'POST'])
@login_required
def edit_entry(entry_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM entries WHERE id = ?", (entry_id,))
    entry = cursor.fetchone()
    if not entry:
        conn.close()
        flash('Entrada não encontrada.', 'danger')
        return redirect(url_for('entradas'))
    if request.method == 'POST':
        entry_date = request.form['entry_date']
        ssd_type = request.form['ssd_type']
        brand = request.form['brand'].strip()
        size = request.form['size'].strip() or None
        quantity = int(request.form['quantity'])
        supplier = request.form['supplier'].strip()
        price = float(request.form['price'])
        qty_diff = quantity - entry['quantity']
        cursor.execute("""
            UPDATE entries SET entry_date=?, ssd_type=?, brand=?, size=?, quantity=?, supplier=?, price=? WHERE id=?
        """, (entry_date, ssd_type, brand, size, quantity, supplier, price, entry_id))
        # Update product aggregate
        cursor.execute("""
            SELECT * FROM products WHERE ssd_type=? AND brand=? AND (size = ? OR (size IS NULL AND ? IS NULL))
        """, (ssd_type, brand, size, size))
        product = cursor.fetchone()
        if product:
            cursor.execute("UPDATE products SET quantity = quantity + ? WHERE id = ?", (qty_diff, product['id']))
        else:
            cursor.execute("INSERT INTO products (ssd_type, brand, size, quantity, total_exited) VALUES (?,?,?,?,0)", (ssd_type, brand, size, quantity))
        conn.commit()
        flash('Entrada atualizada com sucesso.', 'success')
        conn.close()
        return redirect(url_for('entradas'))
    # GET: render edit form
    cursor.execute("SELECT ssd_type, brand, size, quantity FROM products WHERE quantity > 0 ORDER BY brand, size")
    available_products = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return render_template('edit_entry.html', entry=entry, available_products=available_products)

# Delete Entry
@app.route('/entradas/delete/<int:entry_id>', endpoint='delete_entry', methods=['POST'])
@login_required
def delete_entry(entry_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM entries WHERE id = ?", (entry_id,))
    entry = cursor.fetchone()
    if not entry:
        conn.close()
        flash('Entrada não encontrada.', 'danger')
        return redirect(url_for('entradas'))
    # Adjust product quantity
    cursor.execute("""
        SELECT * FROM products WHERE ssd_type=? AND brand=? AND (size = ? OR (size IS NULL AND ? IS NULL))
    """, (entry['ssd_type'], entry['brand'], entry['size'], entry['size']))
    product = cursor.fetchone()
    if product:
        cursor.execute("UPDATE products SET quantity = quantity - ? WHERE id = ?", (entry['quantity'], product['id']))
    cursor.execute("DELETE FROM entries WHERE id = ?", (entry_id,))
    conn.commit()
    flash('Entrada excluída com sucesso.', 'success')
    conn.close()
    return redirect(url_for('entradas'))

# Edit Exit
@app.route('/saidas/edit/<int:exit_id>', endpoint='edit_exit', methods=['GET', 'POST'])
@login_required
def edit_exit(exit_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM exits WHERE id = ?", (exit_id,))
    exit_rec = cursor.fetchone()
    if not exit_rec:
        conn.close()
        flash('Saída não encontrada.', 'danger')
        return redirect(url_for('saidas'))
    if request.method == 'POST':
        exit_date = request.form['exit_date']
        ssd_type = request.form['ssd_type']
        brand = request.form['brand'].strip()
        size = request.form['size'].strip() or None
        quantity = int(request.form['quantity'])
        supplier = request.form['supplier'].strip()
        client = request.form['client'].strip()
        qty_diff = quantity - exit_rec['quantity']
        # Update exit record
        cursor.execute("""
            UPDATE exits SET exit_date=?, ssd_type=?, brand=?, size=?, quantity=?, supplier=?, client=? WHERE id=?
        """, (exit_date, ssd_type, brand, size, quantity, supplier, client, exit_id))
        # Adjust product aggregate
        cursor.execute("""
            SELECT * FROM products WHERE ssd_type=? AND brand=? AND (size = ? OR (size IS NULL AND ? IS NULL))
        """, (ssd_type, brand, size, size))
        product = cursor.fetchone()
        if product:
            # Update quantity: subtract new quantity, add old quantity -> quantity = quantity - qty_diff
            cursor.execute("UPDATE products SET quantity = quantity - ?, total_exited = total_exited + ? WHERE id = ?", (qty_diff, qty_diff, product['id']))
        else:
            # If product not found, create with negative quantity?
            cursor.execute("INSERT INTO products (ssd_type, brand, size, quantity, total_exited) VALUES (?,?,?, ?, ?)", (ssd_type, brand, size, -quantity, quantity))
        conn.commit()
        flash('Saída atualizada com sucesso.', 'success')
        conn.close()
        return redirect(url_for('saidas'))
    # GET: render edit form
    cursor.execute("SELECT ssd_type, brand, size, quantity FROM products WHERE quantity > 0 ORDER BY brand, size")
    available_products = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return render_template('edit_exit.html', exit=exit_rec, available_products=available_products)

# Delete Exit
@app.route('/saidas/delete/<int:exit_id>', endpoint='delete_exit', methods=['POST'])
@login_required
def delete_exit(exit_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM exits WHERE id = ?", (exit_id,))
    exit_rec = cursor.fetchone()
    if not exit_rec:
        conn.close()
        flash('Saída não encontrada.', 'danger')
        return redirect(url_for('saidas'))
    # Adjust product aggregate: add back quantity and subtract from total_exited
    cursor.execute("""
        SELECT * FROM products WHERE ssd_type=? AND brand=? AND (size = ? OR (size IS NULL AND ? IS NULL))
    """, (exit_rec['ssd_type'], exit_rec['brand'], exit_rec['size'], exit_rec['size']))
    product = cursor.fetchone()
    if product:
        cursor.execute("UPDATE products SET quantity = quantity + ?, total_exited = total_exited - ? WHERE id = ?", (exit_rec['quantity'], exit_rec['quantity'], product['id']))
    cursor.execute("DELETE FROM exits WHERE id = ?", (exit_id,))
    conn.commit()
    flash('Saída excluída com sucesso.', 'success')
    conn.close()
    return redirect(url_for('saidas'))


# API endpoint for AJAX stock checking
@app.route('/api/check_stock')
@login_required
def check_stock():
    ssd_type = request.args.get('ssd_type')
    brand = request.args.get('brand', '').strip()
    size = request.args.get('size', '').strip() or None
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT quantity FROM products 
        WHERE ssd_type = ? AND brand = ? AND (size = ? OR (size IS NULL AND ? IS NULL))
    """, (ssd_type, brand, size, size))
    row = cursor.fetchone()
    conn.close()
    
    qty = row['quantity'] if row else 0
    return jsonify({'quantity': qty})

# Export to Excel Route
@app.route('/export/excel/<report_type>')
@login_required
def export_excel(report_type):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Style variables
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo theme
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    title_font = Font(name="Arial", size=14, bold=True, color="1E293B")
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    if report_type == 'entradas':
        ws.title = "Entradas de Estoque"
        # Title block
        ws['A1'] = "Relatório de Entradas de Estoque - StockFlow"
        ws['A1'].font = title_font
        ws.merge_cells('A1:G1')
        ws['A2'] = f"Gerado em: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A2'].font = Font(name="Arial", size=9, italic=True, color="64748B")
        ws.merge_cells('A2:G2')
        
        # Empty row
        ws.append([])
        
        # Headers
        headers = ["Data Entrada", "Tipo", "Marca", "Tamanho", "Quantidade", "Fornecedor", "Preço (R$)"]
        ws.append(headers)
        
        cursor.execute("SELECT entry_date, ssd_type, brand, size, quantity, supplier, price FROM entries ORDER BY entry_date DESC")
        rows = cursor.fetchall()
        for row in rows:
            ws.append([
                row['entry_date'], 
                row['ssd_type'], 
                row['brand'], 
                row['size'] if row['size'] else '-', 
                row['quantity'], 
                row['supplier'], 
                row['price']
            ])
            
    elif report_type == 'saidas':
        ws.title = "Saídas de Estoque"
        # Title block
        ws['A1'] = "Relatório de Saídas de Estoque - StockFlow"
        ws['A1'].font = title_font
        ws.merge_cells('A1:G1')
        ws['A2'] = f"Gerado em: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A2'].font = Font(name="Arial", size=9, italic=True, color="64748B")
        ws.merge_cells('A2:G2')
        
        # Empty row
        ws.append([])
        
        # Headers
        headers = ["Data Saída", "Tipo", "Marca", "Tamanho", "Quantidade", "Fornecedor", "Cliente"]
        ws.append(headers)
        
        cursor.execute("SELECT exit_date, ssd_type, brand, size, quantity, supplier, client FROM exits ORDER BY exit_date DESC")
        rows = cursor.fetchall()
        for row in rows:
            ws.append([
                row['exit_date'], 
                row['ssd_type'], 
                row['brand'], 
                row['size'] if row['size'] else '-', 
                row['quantity'], 
                row['supplier'], 
                row['client']
            ])
    else:
        conn.close()
        return "Tipo de relatório inválido", 400
        
    conn.close()
    
    # Apply formatting
    # Format Headers (row 4)
    for col_idx in range(1, 8):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Format Data rows
    for r_idx in range(5, ws.max_row + 1):
        for c_idx in range(1, 8):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = thin_border
            
            # Alignments
            if c_idx in [1, 2, 4, 5]: # Date, Type, Size, Qty
                cell.alignment = Alignment(horizontal="center")
            elif c_idx == 7 and report_type == 'entradas': # Price
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right")
                
    # Auto-adjust column width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Adjust title height
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[4].height = 24
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"{report_type}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(out, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Export to PDF Route
@app.route('/export/pdf/<report_type>')
@login_required
def export_pdf(report_type):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    
    styles = getSampleStyleSheet()
    
    # Custom Styles
    title_style = ParagraphStyle(
        name='TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#1E293B'),
        spaceAfter=6
    )
    
    subtitle_style = ParagraphStyle(
        name='SubTitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=20
    )
    
    header_style = ParagraphStyle(
        name='TableHeader',
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1 # Center
    )
    
    cell_style = ParagraphStyle(
        name='TableCell',
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#334155'),
        alignment=1 # Center
    )
    
    cell_left_style = ParagraphStyle(
        name='TableCellLeft',
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#334155'),
        alignment=0 # Left
    )
    
    if report_type == 'entradas':
        title = "Relatório de Entradas de Estoque - SSDs"
        story.append(Paragraph(title, title_style))
        story.append(Paragraph(f"Gerado em: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", subtitle_style))
        
        headers = [
            Paragraph("Data", header_style),
            Paragraph("Tipo", header_style),
            Paragraph("Marca", header_style),
            Paragraph("Tamanho", header_style),
            Paragraph("Qtd", header_style),
            Paragraph("Fornecedor", header_style),
            Paragraph("Preço", header_style)
        ]
        
        cursor.execute("SELECT entry_date, ssd_type, brand, size, quantity, supplier, price FROM entries ORDER BY entry_date DESC")
        rows = cursor.fetchall()
        
        data = [headers]
        for row in rows:
            data.append([
                Paragraph(datetime.datetime.strptime(row['entry_date'], '%Y-%m-%d').strftime('%d/%m/%Y'), cell_style),
                Paragraph(row['ssd_type'], cell_style),
                Paragraph(row['brand'], cell_left_style),
                Paragraph(row['size'] if row['size'] else '-', cell_style),
                Paragraph(str(row['quantity']), cell_style),
                Paragraph(row['supplier'], cell_left_style),
                Paragraph(f"R$ {row['price']:.2f}", cell_style)
            ])
            
        # Page size is Letter (612 x 792 points). Margins: 36 x 2. Available width = 540
        # Columns widths sum to 540
        col_widths = [65, 55, 90, 60, 40, 150, 80]
        
    elif report_type == 'saidas':
        title = "Relatório de Saídas de Estoque - SSDs"
        story.append(Paragraph(title, title_style))
        story.append(Paragraph(f"Gerado em: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", subtitle_style))
        
        headers = [
            Paragraph("Data", header_style),
            Paragraph("Tipo", header_style),
            Paragraph("Marca", header_style),
            Paragraph("Tamanho", header_style),
            Paragraph("Qtd", header_style),
            Paragraph("Fornecedor", header_style),
            Paragraph("Cliente", header_style)
        ]
        
        cursor.execute("SELECT exit_date, ssd_type, brand, size, quantity, supplier, client FROM exits ORDER BY exit_date DESC")
        rows = cursor.fetchall()
        
        data = [headers]
        for row in rows:
            data.append([
                Paragraph(datetime.datetime.strptime(row['exit_date'], '%Y-%m-%d').strftime('%d/%m/%Y'), cell_style),
                Paragraph(row['ssd_type'], cell_style),
                Paragraph(row['brand'], cell_left_style),
                Paragraph(row['size'] if row['size'] else '-', cell_style),
                Paragraph(str(row['quantity']), cell_style),
                Paragraph(row['supplier'], cell_left_style),
                Paragraph(row['client'], cell_left_style)
            ])
            
        col_widths = [65, 55, 90, 60, 40, 115, 115]
    else:
        conn.close()
        return "Tipo de relatório inválido", 400
        
    conn.close()
    
    # Build styled ReportLab table
    t = Table(data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')])
    ]))
    
    story.append(t)
    doc.build(story)
    
    pdf_buffer.seek(0)
    filename = f"{report_type}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(pdf_buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")

if __name__ == '__main__':
    app.run(debug=True, port=5000)
